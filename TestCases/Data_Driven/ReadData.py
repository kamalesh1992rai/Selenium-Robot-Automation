import openpyxl
wk=openpyxl.load_workbook("D://TestData100.xlsx")

def  fetch_number_of_rows(Sheetname):
    sh=wk(Sheetname)
    return sh.max_row


def  fetch_cell_data(Sheetname, row, cell):
    sh=wk(Sheetname)
    cell =sh.cell(row,cell)
    return cell.value

print(fetch_number_of_rows("Sheet1"))
print(fetch_cell_data("Sheet1",1,1))
    

sh=wk["Sheet1"]
print(sh.max_row)
cell =sh.cell(1,1)
print(cell.value)
